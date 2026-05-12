from pathlib import Path
from urllib.request import Request, urlopen
import csv
import re
import sys

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "docs" / "data"
DATA.mkdir(parents=True, exist_ok=True)

URL = "https://files.worldhappiness.report/WHR26_Data_Figure_2.1.xlsx"
XLSX = DATA / "WHR26_Data_Figure_2.1.xlsx"
OUT = DATA / "whr_factors.csv"

def clean(value):
    value = str(value or "").lower()
    value = value.replace("&", " and ")
    value = re.sub(r"[’']", "", value)
    value = re.sub(r"[^a-z0-9+]+", " ", value)
    value = re.sub(r"\bthe\b", "", value)
    return re.sub(r"\s+", " ", value).strip()

ALIASES = {
    "united states of america": "united states",
    "usa": "united states",
    "united kingdom": "united kingdom",
    "russia": "russian federation",
    "south korea": "south korea",
    "korea republic of": "south korea",
    "czech republic": "czechia",
    "vietnam": "viet nam",
    "iran islamic republic of": "iran",
    "syrian arab republic": "syria",
    "laos": "lao pdr",
    "democratic republic of congo": "congo kinshasa",
    "congo democratic republic of": "congo kinshasa",
    "republic of congo": "congo brazzaville",
    "ivory coast": "cote divoire",
    "hong kong": "hong kong sar of china",
    "taiwan": "taiwan province of china",
    "palestine": "state of palestine",
    "swaziland": "eswatini",
}

def key_name(value):
    k = clean(value)
    return ALIASES.get(k, k)

def to_float(value):
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None

def download_if_missing():
    if XLSX.exists() and XLSX.stat().st_size > 10_000:
        print(f"Using existing workbook: {XLSX}")
        return

    print(f"Downloading {URL}")
    req = Request(
        URL,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
        },
    )
    try:
        with urlopen(req, timeout=60) as response:
            XLSX.write_bytes(response.read())
    except Exception as exc:
        raise SystemExit(
            f"Could not download WHR workbook automatically: {exc}\n"
            f"Download it manually from the WHR data-sharing page and save it as:\n"
            f"{XLSX}"
        )

def find_col(headers, groups):
    cleaned = [clean(h) for h in headers]
    for terms in groups:
        terms = [clean(t) for t in terms]
        for i, h in enumerate(cleaned):
            if all(t in h for t in terms):
                return i
    return None

def parse_workbook():
    wb = load_workbook(XLSX, data_only=True, read_only=True)

    best_rows = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))

        for header_idx, row in enumerate(rows[:80]):
            headers = list(row)
            joined = " | ".join(clean(h) for h in headers)

            if "country" not in joined:
                continue

            idx = {
                "country": find_col(headers, [["country", "name"], ["country"]]),
                "score": find_col(headers, [["ladder", "score"], ["life", "evaluation"], ["happiness", "score"]]),
                "gdp": find_col(headers, [["explained", "gdp"], ["gdp", "per", "capita"], ["log", "gdp"]]),
                "social": find_col(headers, [["explained", "social"], ["social", "support"]]),
                "life": find_col(headers, [["explained", "healthy"], ["healthy", "life"]]),
                "freedom": find_col(headers, [["explained", "freedom"], ["freedom", "life"]]),
                "generosity": find_col(headers, [["explained", "generosity"], ["generosity"]]),
                "corruption": find_col(headers, [["explained", "corruption"], ["perceptions", "corruption"], ["corruption"]]),
                "residual": find_col(headers, [["dystopia"], ["residual"]]),
            }

            if idx["country"] is None:
                continue

            parsed = []
            for data_row in rows[header_idx + 1:]:
                country = data_row[idx["country"]] if idx["country"] < len(data_row) else None
                if not country or not str(country).strip():
                    continue

                item = {
                    "country_key": key_name(country),
                    "country": str(country).strip(),
                    "score": to_float(data_row[idx["score"]]) if idx["score"] is not None and idx["score"] < len(data_row) else None,
                    "gdp": to_float(data_row[idx["gdp"]]) if idx["gdp"] is not None and idx["gdp"] < len(data_row) else None,
                    "social": to_float(data_row[idx["social"]]) if idx["social"] is not None and idx["social"] < len(data_row) else None,
                    "life": to_float(data_row[idx["life"]]) if idx["life"] is not None and idx["life"] < len(data_row) else None,
                    "freedom": to_float(data_row[idx["freedom"]]) if idx["freedom"] is not None and idx["freedom"] < len(data_row) else None,
                    "generosity": to_float(data_row[idx["generosity"]]) if idx["generosity"] is not None and idx["generosity"] < len(data_row) else None,
                    "corruption": to_float(data_row[idx["corruption"]]) if idx["corruption"] is not None and idx["corruption"] < len(data_row) else None,
                    "residual": to_float(data_row[idx["residual"]]) if idx["residual"] is not None and idx["residual"] < len(data_row) else None,
                }

                has_factor = any(item[k] is not None for k in ["gdp", "social", "life", "freedom", "generosity", "corruption", "residual"])
                if has_factor:
                    parsed.append(item)

            if len(parsed) > len(best_rows):
                best_rows = parsed

    if not best_rows:
        raise SystemExit("Could not find WHR factor columns in the workbook.")

    return best_rows

def write_csv(rows):
    fields = ["country_key", "country", "score", "gdp", "social", "life", "freedom", "generosity", "corruption", "residual"]
    with OUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    print(f"Wrote {len(rows)} countries to {OUT}")

def main():
    download_if_missing()
    rows = parse_workbook()
    write_csv(rows)

if __name__ == "__main__":
    main()
